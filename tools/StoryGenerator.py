import json
import openpyxl



#create a main function that will be run via command line
def main():
    #open the excel file
    wb=None
    try:
        wb = openpyxl.load_workbook('../project-CB-Story.xlsx')
    except:
        print("Error: Cannot open the excel file")
        return

    #open the sheet
    sheet = wb['Sheet1']
    #create a dictionary to store the data
    data = {'storypoints': []}
    index=0
    #iterate over the rows in the sheet
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4, values_only=True):
        #extract the data from the row
        game_time, assigned_channel, story_text,sound_file_name = row
        print(f"time: {game_time} channel: {assigned_channel} story: {story_text} sound: {sound_file_name}")
        #store the data in the dictionary
        data['storypoints'].append({"gameTime": str(game_time), "assignedChannel": assigned_channel, "text": story_text, "soundFileName": sound_file_name})
        index+=1
    #write the data to a JSON file
    json_data = json.dumps(data, indent=4)
    print(json_data)

    with open('../source/Story/StoryPointTest.json', 'w') as f:
        f.write(json_data)

if __name__ == "__main__":
    main()