#might use https://github.com/neonbjb/tortoise-tts to replace google
from gtts import gTTS
from pydub import AudioSegment


import openpyxl
import os
import shutil

def main():
    wb=None
    try:
        wb = openpyxl.load_workbook('../project-CB-Story.xlsx')
    except:
        print("Error: Cannot open the excel file")
        return
    #temp path
    tempPathRoot=os.path.join(os.getcwd(),'temp')
    tempPath=os.path.join(tempPathRoot,'Sound')
    os.makedirs(tempPath, exist_ok=True)
    #open the sheet
    sheet = wb['index']
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4, values_only=True):
        game_time, assigned_channel, story_text,sound_file_name = row
        print(f"time: {game_time} channel: {assigned_channel} story: {story_text} sound: {sound_file_name}")
        tts = gTTS(text=story_text, lang='en')
        
        # create temp folder in the root directory
        
        tts.save(f'{tempPath}/{sound_file_name}.mp3')
        print(f"Sound file {sound_file_name} generated")

        #convert the mp3 file to wav
        sound = AudioSegment.from_mp3(f'{tempPath}/{sound_file_name}.mp3')
        sound.export(f'../source/sound/{sound_file_name}.wav', format="wav")
        
    #delete the temp folder and all files in it
    shutil.rmtree(tempPathRoot)


if __name__ == "__main__":
    main()